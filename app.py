import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl

# --- Configuración de la Página ---
st.set_page_config(page_title="Cotizador v11Base", layout="wide")
st.title("Gestor de Cotizaciones - v11Base")

# --- Lógica de Negocio (El Corazón del v11) ---
def calcular_valor_final(pais, moneda, costo, trm):
    """
    Aplica la lógica del v11Base:
    1. Si es Ecuador: El costo pasa directo (Excepción).
    2. Si es USD (y no Ecuador): Se divide por la TRM.
    3. Si es Local: Pasa directo.
    """
    if pais == "Ecuador":
        return costo
    elif moneda == "USD":
        if trm and trm > 0:
            return costo / trm
        return 0
    else:
        return costo

# --- Interfaz de Usuario ---
col1, col2 = st.columns([1, 2])

with col1:
    st.subheader("1. Configuración")
    uploaded_file = st.file_uploader("Cargar archivo v11Base (.xlsx / .xlsm)", type=["xlsx", "xlsm"])
    
    st.markdown("---")
    st.subheader("2. Datos de la Cotización")
    fecha_inicio = st.date_input("Fecha Inicio")
    fecha_fin = st.date_input("Fecha Fin")
    pais = st.selectbox("País", ["Colombia", "Ecuador", "Perú", "México", "Chile", "Argentina"])
    
    moneda = st.radio("Moneda", ["Local", "USD"], horizontal=True)
    costo = st.number_input("Costo (Valor)", min_value=0.0, format="%.2f")
    
    # TRM solo es relevante si es USD, pero siempre la pedimos por si acaso
    trm = st.number_input("Tasa de Cambio (TRM)", value=1.0, min_value=0.0001, format="%.2f")

with col2:
    st.subheader("3. Vista Previa y Procesamiento")
    
    if uploaded_file is not None:
        try:
            # Leemos el archivo para mostrar info básica (usando openpyxl para no romper macros)
            wb = openpyxl.load_workbook(uploaded_file)
            sheet_names = wb.sheetnames
            st.success(f"Archivo cargado: {uploaded_file.name}")
            st.write(f"Hojas detectadas: {', '.join(sheet_names)}")
            
            # Botón para procesar
            if st.button("Insertar Cotización en v11Base", type="primary"):
                
                # 1. Calcular el valor a insertar
                valor_a_insertar = calcular_valor_final(pais, moneda, costo, trm)
                
                # 2. Seleccionar la hoja (Asumimos 'INPUT cost' por tu historial, o la primera si no existe)
                target_sheet_name = "INPUT cost"
                if target_sheet_name not in wb.sheetnames:
                    st.warning(f"No encontré la hoja '{target_sheet_name}', usaré la primera activa.")
                    ws = wb.active
                else:
                    ws = wb[target_sheet_name]
                
                # 3. Encontrar la siguiente fila vacía (o insertar línea específica si tienes una regla fija)
                # Aquí agregamos al final como ejemplo seguro
                next_row = ws.max_row + 1
                
                # 4. Escribir datos (Mapeo básico, ajústalo a tus columnas reales del v11)
                # Ejemplo: A=Fecha, B=Pais, C=Valor Final
                ws.cell(row=next_row, column=1, value=fecha_inicio)
                ws.cell(row=next_row, column=2, value=pais)
                ws.cell(row=next_row, column=3, value=valor_a_insertar) 
                
                # Feedback visual
                st.write(f"✅ Dato insertado en fila {next_row}")
                st.write(f"💰 Valor calculado aplicado: {valor_a_insertar:,.2f} (Lógica: {pais}/{moneda})")

                # 5. Guardar en memoria para descargar
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    label="Descargar v11Base Actualizado",
                    data=output,
                    file_name=f"v11Base_Actualizado_{pais}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
        except Exception as e:
            st.error(f"Error procesando el archivo: {e}")
    else:
        st.info("👈 Por favor carga el archivo v11Base para comenzar.")

# --- Debug / Validación ---
st.divider()
valor_test = calcular_valor_final(pais, moneda, costo, trm)
st.caption(f"Validación de Lógica en tiempo real: Si insertaras ahora, el valor sería {valor_test:,.2f}")
